# Modifying heatmappage.py to add in a new button for spreadsheet transformations
# Sydney : Modigying stuff to add a button to add a moon scrape page.
# Sydney : Modigying stuff to add a button to add a moon scrape page.
#!/usr/bin/python
# -*- coding: utf-8 -*-
from datetime import datetime
from tkinter.ttk import Style

import matplotlib
import pandas as pd
import numpy as np

import scipy.spatial as ss
import PIL.Image
from PIL import ImageTk
import math
import shutil

from errors import *
import time

import platform

from heatmap import HeatMapOptionsBox
from heatmap import InputError
from heatmappage import HeatMapPage
from heatmappage import StartPage
from kde import KDE_Page, KDE_Calculation_Page
from joins_home import Joins_Home_Page
from joins_light import Joins_Page_Light
from joins_rubbing import Joins_Page_Rubbing
from transformations import Transformations_Page
from moon_scrape_home import Moon_Scrape_Home_Page
from moon_scrape_doc_to_excel import Doc_To_Excel_Moon_Scrape_Page
from moon_scrape_excel_to_excel import Excel_To_Excel_Moon_Scrape_Page
from moon_scrape_excel_to_sheet import Excel_To_Sheet_Moon_Scrape_Page
from categories import Categories_Page
from joins_both import Joins_Page_Both
import matplotlib.cm as cm
import matplotlib.image as mpimg

import time
import threading

matplotlib.use('TkAgg')

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from mpl_toolkits.mplot3d import Axes3D
import matplotlib.axes
import PIL.ImageTk
from tkinter import *
import tkinter as tk
from tkinter import ttk
from tkinter import filedialog
from tkinter import messagebox
import openpyxl
from openpyxl import load_workbook, Workbook

import os
import json
import re
import shlex

from tksheet import Sheet
from pages import PageOne, PageTwo, PageThree
import csv

LARGE_FONT = ('Bell Gothic Std Black', 40, 'bold')
MEDIUM_FONT = ('Bell Gothic Std Black', 25, 'bold')
BUTTON_FONT = ('Calibiri', 14, 'bold')
BACKGROUND_COLOR = '#407297'
LIGHT_BLUE = '#d4e1fa'


def errorMessage(e):
    messagebox.showerror('Error', e.message)


class ZooMapper(tk.Tk):

    """
    Main class of the application
    Creates the tk window, handles importing a spreadsheet, and switching pages.
    """

    def __init__(self, *args, **kwargs):
        tk.Tk.__init__(self, *args, **kwargs)

        # tk.Tk.iconbitmap(self, default="resources/clienticon.ico")

        tk.Tk.wm_title(self, 'Zoo Mapper')
        tk.Tk.wm_geometry(self, '1440x810')

        self.protocol('WM_DELETE_WINDOW', lambda : \
                      self.close_application())

        # Opens in maximized window on Windows and Mac
        # Opens in fullscreen for Linux

        os_name = platform.system()
        if os_name == 'Windows':
            self.state('zoomed')
            self.isFullScreen = False
        elif os_name == 'Darwin':
            self.attributes('-fullscreen', True)
            self.isFullScreen = False
        else:
            self.attributes('-fullscreen', True)
            self.isFullScreen = True

        # Escape exits fullscreen

        self.bind('<Escape>', self.toggle_fullscreen)

        self.minsize(400, 300)

        menu = Menu(self)
        tk.Tk.config(self, menu=menu)

        # Define Menus

        file_menu = Menu(menu, tearoff=0)
        edit_menu = Menu(menu, tearoff=0)
        view_menu = Menu(menu, tearoff=0)
        about_menu = Menu(menu, tearoff=0)
        menu.add_cascade(label='File', menu=file_menu)
        menu.add_cascade(label='Edit', menu=edit_menu)
        menu.add_cascade(label='View', menu=view_menu)
        menu.add_cascade(label='About', menu=about_menu)

        # File Menu Options

        file_menu.add_command(label='Import Spreadsheet',
                              command=self.get_spreadsheet)
        file_menu.add_command(label='Reduce Spreadsheet',
                              command=self.reduce_spreadsheet)

        # file_menu.add_command(label="Import Habitat", command=self.get_image)

        # Edit Menu Options

        # View Menu Options
        # view_menu.add_command(label="Display Coordinates", command=self.get_list)
        # view_menu.add_command(label="Hide Coordinates", command=self.remove_list)

        # About Menu Options

        view_menu.add_command(label='Toggle Fullscreen',
                              command=self.toggle_fullscreen)
        about_menu.add_command(label='Developers',
                               command=self.print_dev)

        container = tk.Frame(self)
        container.pack(side='top', fill='both', expand=True)
        container.grid_rowconfigure(0, weight=1)
        container.grid_columnconfigure(0, weight=1)

        self.container_please = container
        self.spreadsheet_path = ''

        self.frames = {}

        self.df = []

        self.sheet_index = 0
        self.sheet_name = ''
        self.selected_sheet = False
        self.submit_selection = False

        #Took out """KDE_Page"""
        for F in (StartPage, PageOne, PageTwo, HeatMapPage, Transformations_Page, 
                  Categories_Page, Joins_Home_Page, Joins_Page_Both, Joins_Page_Light, Joins_Page_Rubbing, 
                  Moon_Scrape_Home_Page, Excel_To_Excel_Moon_Scrape_Page, Excel_To_Sheet_Moon_Scrape_Page, Doc_To_Excel_Moon_Scrape_Page):
            frame = F(container, self)
            frame.config(bg=BACKGROUND_COLOR)
            self.frames[F] = frame

            frame.grid(row=0, column=0, sticky=NSEW)
        
        self.show_frame(StartPage)

    def close_application(self):
        """
        Method called when the app is closed. Cloases any open spreadsheets.
        """

        if hasattr(self, 'sheet'):
            self.sheet.close(self)
        self.destroy()

    def popup_sheet_selection(self, sheet_names):

        popup = tk.Toplevel(self)
        popup.wm_title('Select A Sheet Name')
        label = ttk.Label(popup,
                          text='Please select a sheet name containing your data, then reopen this file.'
                          )
        label.pack(side='top', fill='x', pady=10)

        options = sheet_names

        cl = StringVar(popup)
        cl.set('Select a sheet name')

        dr = OptionMenu(popup, cl, *options)
        dr.pack()

        B1 = Button(popup, text='Cancel', command=popup.destroy)
        B1.pack(side='bottom', fill='y')

        B2 = Button(popup, text='Submit', command=lambda : \
                    self.submit_sheet_choice(popup, sheet_names,
                    cl.get()))
        B2.pack(side='bottom', fill='y')

        return self.sheet_index

    # In fullscreen, pressing escape will exit fullscreen

    def toggle_fullscreen(self, event=None):
        """
        In fullscreen, pressing escape will exit fullscreen
        """

        if self.isFullScreen:
            self.attributes('-fullscreen', False)
            self.isFullScreen = False
        else:
            self.attributes('-fullscreen', True)
            self.isFullScreen = True

    def get_plot_creation_options(self, data_frame, saved_import=None):
        """
        Gets options from import window and does a quick error check for required columns
        """

        plot_options = {'heatmap_options': {}}

        validParameters = False
        while validParameters == False:
            try:

                # if self.sheet_choice == "":
                # ....break

                heat_options_box = HeatMapOptionsBox(
                    StartPage,
                    data_frame,
                    (plot_options, 'heatmap_options'),
                    self.spreadsheet_path,
                    self.sheet_choice,
                    saved_defaults=saved_import,
                    )
                heat_options_box.wait_window(heat_options_box)
                options = plot_options['heatmap_options']
                saved_import = options

                if options['x_column'] == '':
                    errorMessage(Error.NOXCOL)
                elif options['y_column'] == '':
                    errorMessage(Error.NOYCOL)
                elif options['name_column'] == '':
                    noName = messagebox.askyesno(title=None,
                            message='Are you sure you do not want to choose a Name Column?'
                            )
                    if noName:
                        validParameters = True
                else:
                    validParameters = True
            except KeyError:

                validParameters = True
                options = None

        return options

    def get_spreadsheet(self, saved_import=None):
        """
        Opens file explorer to select a spreadhsset and creates a dataframe from selected spreadsheet.
        If loading a saved import, will instead open spreadsheet from import
        """

        if hasattr(self, 'sheet'):
            self.sheet.close(self)

        validFile = False
        self.sheet_choice = ''

        while validFile == False:
            try:

                if saved_import == None:
                    filename = filedialog.askopenfilename(initialdir=''
                            , title='Select a File',
                            filetypes=(('Excel files', '*.xlsx*'),
                            ('CSV files', '*.csv*'), ('all files', '*.*'
                            )))
                else:
                    filename = saved_import['spreadsheet_path']
                    self.sheet_choice = saved_import['sheet_name']

                self.spreadsheet_path = filename

                file_type = filename[filename.index('.'):]

                data = None

                # Here the file type choice is accounted for, as well as the error of an incorrect file choice but it is not handled, we need to do this.

                if file_type == '.xlsx':

                    data = pd.read_excel(filename, sheet_name=None)
                    xl_data = pd.ExcelFile(filename)

                    if len(xl_data.sheet_names) > 1:

                        # data.parse(self.sheet_choice)
                        # parent.sheet_name = self.sheet_choice

                        if saved_import == None:
                            popup = tk.Toplevel(self)
                            choice_window = Sheet_Entry(self, popup,
                                    xl_data.sheet_names)
                            choice_window.dr.wait_window(choice_window.dr)

                        if self.sheet_choice != '':
                            data = pd.read_excel(filename,
                                    sheet_name=self.sheet_choice)

                        if self.sheet_choice is not None:
                            self.selected_sheet = True
                            validFile = True
                    else:
                        data = pd.read_excel(filename, sheet_name=0)
                        validFile = True
                        self.selected_sheet = True
                        self.submit_selection = True
                        self.sheet_choice = '0'

                    if self.selected_sheet:

                        # data = pd.read_excel(filename, sheet_name=self.sheet_index)

                        validFile = True
                elif file_type == '.csv':
                    data = pd.read_csv(filename)
                    validFile = True
                    self.selected_sheet = True
                else:
                    errorMessage(Error.FILETYPE)
                    break

                if self.selected_sheet and self.sheet_choice != '':

                    global heatmap_options

                    heatmap_options = \
                        self.get_plot_creation_options(data,
                            saved_import=saved_import)
                    self.df = data

                    if heatmap_options != None:
                        global heat_frame
                        heat_frame = HeatMapPage(self.container_please,
                                self, data, heatmap_options)
                        heat_frame.config(bg=LIGHT_BLUE)
                        heat_frame.grid(row=0, column=0, sticky='nsew')
                        self.frames[HeatMapPage] = heat_frame
            except ValueError:

                break
                
    def reduce_spreadsheet(self, saved_import=None):
        """
        Opens file explorer to select a spreadsheet and creates a copy where each SessionID will have
        the same number of entries.
        """

        if hasattr(self, 'sheet'):
            self.sheet.close(self)

        validFile = False
        self.sheet_choice = ''

        while validFile == False:
            try:

                if saved_import == None:
                    filename = filedialog.askopenfilename(initialdir=''
                            , title='Select a File',
                            filetypes=(('Excel files', '*.xlsx*'),
                            ('CSV files', '*.csv*'), ('all files', '*.*'
                            )))
                else:
                    filename = saved_import['spreadsheet_path']
                    self.sheet_choice = saved_import['sheet_name']

                self.spreadsheet_path = filename

                file_type = filename[filename.index('.'):]

                data = None

                temp_filename = filename
                temp_filename.replace('.xlsx', '')
                temp_filename = temp_filename + '_Reduced.xlsx'
                shutil.copyfile(filename, temp_filename)

                wb = Workbook()
                wb = openpyxl.load_workbook(temp_filename)
                wb_active = wb.active

                total_min = 99
                record_counter = 0
                row_number = 0
                temp_cell = 0

                # Gets the lowest number for records at or above 3.
                for row in wb_active:
                    row_number = row_number + 1
                    if row_number != 1:
                        for cell in row:
                            if cell.value == temp_cell:
                                record_counter = record_counter + 1
                            if cell.value != temp_cell:
                                if record_counter >= 3:
                                    if record_counter < total_min:
                                        total_min = record_counter
                                        record_counter = 0
                            temp_cell = cell.value
                            break

                delete_list = []
                row_number = 0
                record_counter = 0
                for row in wb_active:
                    row_number = row_number + 1
                    if row_number != 1:
                        for cell in row:
                            if cell.value == temp_cell:
                                record_counter = record_counter + 1
                            if cell.value != temp_cell:
                                record_counter = 0
                            if record_counter >= total_min:
                                # watch for errors here
                                delete_list.append(row_number)
                            temp_cell = cell.value
                            break
                safety_counter = 0
                while len(delete_list) > 0:
                    temp_length = len(delete_list)
                    wb_active.delete_rows(delete_list[temp_length - 1])
                    del delete_list[temp_length - 1]
                    safety_counter = safety_counter + 1
                    if safety_counter >= 99999:
                        break

                wb.save(temp_filename)



                # Here the file type choice is accounted for, as well as the error of an incorrect file choice but it is not handled, we need to do this.

                if file_type == '.xlsx':

                    data = pd.read_excel(temp_filename, sheet_name=None)
                    xl_data = pd.ExcelFile(temp_filename)

                    if len(xl_data.sheet_names) > 1:

                        # data.parse(self.sheet_choice)
                        # parent.sheet_name = self.sheet_choice

                        if saved_import == None:
                            popup = tk.Toplevel(self)
                            choice_window = Sheet_Entry(self, popup,
                                    xl_data.sheet_names)
                            choice_window.dr.wait_window(choice_window.dr)

                        if self.sheet_choice != '':
                            data = pd.read_excel(temp_filename,
                                    sheet_name=self.sheet_choice)

                        if self.sheet_choice is not None:
                            self.selected_sheet = True
                            validFile = True
                    else:
                        data = pd.read_excel(temp_filename, sheet_name=0)
                        validFile = True
                        self.selected_sheet = True
                        self.submit_selection = True
                        self.sheet_choice = '0'

                    if self.selected_sheet:

                        # data = pd.read_excel(filename, sheet_name=self.sheet_index)

                        validFile = True
                elif file_type == '.csv':
                    data = pd.read_csv(filename)
                    validFile = True
                    self.selected_sheet = True
                else:
                    errorMessage(Error.FILETYPE)
                    break

                if self.selected_sheet and self.sheet_choice != '':

                    global heatmap_options

                    heatmap_options = \
                        self.get_plot_creation_options(data,
                            saved_import=saved_import)
                    self.df = data

                    if heatmap_options != None:
                        global heat_frame
                        heat_frame = HeatMapPage(self.container_please,
                                self, data, heatmap_options)
                        heat_frame.config(bg=LIGHT_BLUE)
                        heat_frame.grid(row=0, column=0, sticky='nsew')
                        self.frames[HeatMapPage] = heat_frame

            except ValueError:


                break

    def update_graph(self, sheet=None, saved_import=None):
        """
        Updates the displayed graph with the newest available dataframe
        """

        if sheet != None:
            data = sheet.get_sheet_data(return_copy=False,
                    get_header=False, get_index=False)
            col = sheet.get_sheet_data(return_copy=True,
                    get_header=True, get_index=True)
            cols = []

            # getting columns

            for i in col:
                if not isinstance(i, list):
                    cols.append(i)
            newdata = pd.DataFrame(data, columns=cols)

            # print(newdata)

            heat_frame = HeatMapPage(self.container_please, self,
                    newdata, heatmap_options)
            heat_frame.config(bg=LIGHT_BLUE)
            heat_frame.grid(row=0, column=0, sticky='nsew')
            self.frames[HeatMapPage] = heat_frame
        else:
            self.show_frame(HeatMapPage)

    def load_import(self, page):
        """
        Handles loading a saved import. Reads the options saved.
        """

        imported_file = {'import_file': {}}
        load_window = LoadWindow(page, (imported_file, 'import_file'))
        load_window.wait_window(load_window)

        import_config_file = imported_file['import_file']

        if import_config_file != None:
            save_file = open('saves/' + import_config_file, 'r')
            saved_options = json.load(save_file)
            save_file.close()

            self.get_spreadsheet(saved_import=saved_options)

    def print_dev(self):
        messagebox.showinfo('Developers',
                            '''Farah Aljishi
Tyler Ballance
Derek Baum
Ryan Bonacquisti
Kevin Bookwalter
Nishant Chintala
Bradley Fusting
Charlie Hannum
Samuel Herring
Debra Lymon
Fanchao Meng
Ntsee Ndingwan
Bobby Stahl
Justin Taylor
Jake Wise
Yuzu Wu
Michael Yacucci
Shuobofang Yang
Arthur Marino
''')

    def show_frame(self, cont):
        """
        Shows frame at specified index
        """

        frame = self.frames[cont]
        frame.tkraise()

    def show_sheet(self, data_frame, opts):
        """
        Shows spreadsheet in addition to graph.
        """

        self.sheet = SheetView(data_frame, opts, self)
        self.sheet.title(self.spreadsheet_path.split('/')[-1])

    def calc_name_distance(
        self,
        parent,
        data_frame,
        options,
        cal_ratio,
        ):
        """
        Opens calculations window to handle exporting distance calculations between animals
        """

        if hasattr(self, 'name_window'):
            self.name_window.deiconify()
        else:
            self.name_window = CalculationInputPage(parent, data_frame,
                    options, cal_ratio)
            self.name_window.wait_window(self.name_window)


class Sheet_Entry:

    def __init__(
        self,
        parent,
        owner,
        sheet_names,
        ):

        self.main_class = parent

        self.master = owner
        self.label = ttk.Label(self.master,
                               text='Please select a sheet name containing your data, then reopen this file.'
                               )
        self.label.pack(side='top', fill='x', pady=10)
        options = sheet_names
        self.cl = StringVar(self.master)
        self.cl.set('Select a sheet name')
        self.dr = OptionMenu(self.master, self.cl, *options)
        self.dr.pack()
        self.B1 = Button(self.master, text='Cancel',
                         command=self.master.destroy)
        self.B1.pack(side='bottom', fill='y')
        self.B2 = Button(self.master, text='Submit', command=lambda : \
                         self.submit_sheet_choice(self.master,
                         sheet_names, self.cl.get()))
        self.B2.pack(side='bottom', fill='y')

    def submit_sheet_choice(
        self,
        parent,
        sheets,
        choice,
        ):

            # print(sheets.index(choice))
            # self.sheet_index = sheets.index(choice)

        if not isinstance(choice, str) or choice == '':
            parent.destroy()

        self.main_class.sheet_choice = str(choice)
        self.main_class.selected_sheet = True
        self.main_class.submit_selection = True

            # self.main_class.get_spreadsheet()

        parent.destroy()


class LoadWindow(tk.Toplevel):

    """
    Class that creates Toplevel window for loading or deleting a saved import.
    """

    def __init__(self, parent, out_dict):
        tk.Toplevel.__init__(self, parent)

        self.protocol('WM_DELETE_WINDOW', lambda arg=out_dict: \
                      self.return_cancel(arg))

        self.title('Load Import')

        tk.Toplevel.wm_geometry(self, '300x400')

        frm = tk.Frame(self, borderwidth=4, relief='raised')
        self.frame = frm
        self.frame.config(bg='white')
        frm.pack(fill='both', expand=True)

        label = tk.Label(frm, text='Select an import to load or delete')
        label.pack(pady=10)

        saves = Listbox(frm, selectmode='browse')
        saves.pack(pady=15)

        saved_imports = self.list_saves_names()
        for (key, value) in saved_imports.items():
            saves.insert(END, key)

        b_import = tk.Button(frm, text='LOAD')
        b_import['command'] = lambda : \
            self.return_file_to_import(saves, out_dict, saved_imports)
        b_import.pack(pady=4)

        b_delete = tk.Button(frm, text='DELETE')
        b_delete['command'] = lambda : self.delete_save(saves)
        b_delete.pack(pady=4)

        b_ok = tk.Button(frm, text='CANCEL')
        b_ok['command'] = lambda : self.return_cancel(out_dict)
        b_ok.pack(pady=4)

    def list_saves_names(self):
        """
        Creates a list of import names
        """

        files = {}
        for file in os.listdir('saves'):
            if file.endswith('.json'):
                file_name = file[:-5]
                files[file_name] = file
        return files

    def return_file_to_import(
        self,
        listbox,
        out_dict,
        saved_imports,
        ):
        """
        Destroys window and returns dict with selected file to import
        """

        (d, key) = out_dict
        selection = str(listbox.get(ACTIVE))
        d[key] = saved_imports[selection]
        self.destroy()

    def return_cancel(self, out_dict):
        """
        Handles when load is cancelled.
        """

        (d, key) = out_dict
        d[key] = None
        self.destroy()

    def delete_save(self, listbox):
        """
        Handles deleting a saved import
        """

        selection = str(listbox.get(ACTIVE))
        try:
            file_index = listbox.curselection()[0]
            os.remove('saves/' + selection + '.json')
            listbox.delete(file_index)
        except IndexError:
            pass


class CalculationInputPage(tk.Toplevel):

    """
    Class for Calculation Output Overlay. Exists as a Toplevel widget on Graph Page
    """

    def __init__(
        self,
        parent,
        data_frame,
        options,
        cal_ratio,
        ):

        # Creates Toplevel widget over parent and initializes Cancel behavior

        tk.Toplevel.__init__(self, parent)
        self.parent = parent
        self.protocol('WM_DELETE_WINDOW', lambda : self.return_cancel())
        self.attributes('-topmost', 'true')

        # Holds passed in name for xcolumn, ycolumn and names in Dataframe

        self.xcol_name = options['x_column']
        self.ycol_name = options['y_column']
        self.name_column = options['name_column']
        self.x_ratio = options['x_ratio']
        if self.x_ratio != '':
            self.x_ratio = float(self.x_ratio)
        self.y_ratio = options['y_ratio']
        if self.y_ratio != '':
            self.y_ratio = float(self.y_ratio)
        self.z_ratio = options['z_ratio']
        if self.z_ratio != '':
            self.z_ratio = float(self.z_ratio)
        self.cal_ratio = cal_ratio
        self.filename = None
        self.outputFile = None
        self.zcol_name = options['z_column']

        # Sets up widget with size, frame and title.

        self.title('Calculation Input')
        tk.Toplevel.wm_geometry(self, '350x400')
        self.data = data_frame.copy()
        frm = tk.Frame(self, borderwidth=4, relief='raised')
        self.frame = frm
        self.frame.config(bg='white')
        frm.pack(fill='both', expand=True)

        # Creates first label and option dropdown for the Time Column

        label = tk.Label(frm, text='Choose Time Element')
        label.pack(pady=10)
        self.columns_list = data_frame.columns

        self.time_column_var = tk.StringVar()
        time_filter_list = HeatMapOptionsBox.filterOptions(self,
                [['Date'], ['Time']], self.columns_list)

        if 'Date' in time_filter_list and 'Time' in time_filter_list:
            time_filter_list.append('Date + Time')

        self.time_column_dropdown = tk.OptionMenu(frm,
                self.time_column_var, *time_filter_list)
        self.time_column_dropdown.pack()

        def check_date_time(*args):
            """
            Checks if the column is only Date or Time and displays a warning
            """

            if self.time_column_var.get() in ['Date', 'Time']:
                self.time_warning_label.config(text='Choosing '
                        + str(self.time_column_var.get())
                        + ' will calculate only one distance per '
                        + str(self.time_column_var.get()) + '.')
            else:
                self.time_warning_label.config(text='')

        self.time_column_var.trace('w', check_date_time)

        self.time_warning_label = tk.Label(frm, text='', bg='white',
                fg='red')
        self.time_warning_label.pack(pady=4)

        # Creates label and Entry field for List of Names. Currently only works with 2 names with a space in the middle

        self.name_label = tk.Label(frm, text='Select 2 Names to Calculate Distances')
        self.name_label.pack(pady=4)
        #oringal implementation of text box
        #self.name_var = tk.Entry(frm)
        
        names = data_frame[options['name_column']].unique()
        names = list(filter(None, names))
        print('names:', names)
        self.name_var = tk.Listbox(frm, selectmode=tk.MULTIPLE, height=4)
        values = names[1:]
        for val  in values:
            self.name_var.insert(tk.END, val)
        self.name_var.pack()
        
        self.name_var.pack(pady=4)

        self.b_choose_file = tk.Button(frm, text='Save File As')
        self.b_choose_file['command'] = lambda : self.save_file()
        self.b_choose_file.pack(pady=4)

        self.filename_label = tk.Label(frm)
        self.filename_label.pack()

        label = tk.Label(frm, bg='white',
                         text='''When submitted, calculations will be performed 
 in the background. You can continue using ZooMapper 
 and will be notified when the calculations are complete.''')
        label.pack(pady=10)

        # Creates submit button that runs calculate distances and outputs using the inputs

        self.submit = tk.Button(frm, text='Calculate')
        self.submit['command'] = lambda : \
            self.calculateDistances(self.data,
                                    self.time_column_var.get(),
                                    (self.name_var.get(0) + ' ' + self.name_var.get(1)))
        # debugging lambda to get names
        '''self.submit['command'] = lambda : \
            print((self.name_var.get(self.name_var.curselection()[0]) + ' ' + self.name_var.get(self.name_var.curselection()[1])))
        '''
        # TODO() restrict more than two options
        self.submit.pack(pady=4)

    def save_file(self):
        """
        Opens a file explorer to choose location and name of output file
        """

        self.attributes('-topmost', 'false')
        csv_save = [('CSV Files', '*.csv')]
        outFile = filedialog.asksaveasfile(mode='w',
                filetypes=csv_save, defaultextension=csv_save)
        if outFile != None:
            self.outputFile = outFile
            self.filename = os.path.basename(self.outputFile.name)
            self.filename_label.config(text=self.filename)
        self.attributes('-topmost', 'true')

    def return_cancel(self):
        """
        Cancel function for closing input window
        """

        self.destroy()
        delattr(self.parent, 'name_window')

    def calculateDistances(
        self,
        data_frame,
        time_column,
        names,
        ):
        """
        Checks for errors before creating thread to perform calculations.
        Thread allows for user to continue using application while calculations are performed.
        """

        errCode = []
        if not time_column:
            errCode.append('A')
        if len(names) < 2:
            errCode.append('B')
        if not self.outputFile:
            errCode.append('C')
        if len(errCode) > 0:
            errMsg = ''
            if 'A' in errCode:
                errMsg = errMsg + 'Time Element is not specified.\n'
            if 'B' in errCode:
                errMsg = errMsg \
                    + 'Not enough names are specified (minimum 2 required).\n'
            if 'C' in errCode:
                errMsg = errMsg + 'No output file specified.\n'
            self.attributes('-topmost', 'false')
            self.wait_window(InputError(self, errMsg.rstrip('\n')))
            self.attributes('-topmost', 'true')
        else:
            thread = \
                threading.Thread(target=self.calculateDistances_thread,
                                 args=(data_frame, time_column, names))
            thread.start()

            # Destroys window when complete

            self.destroy()
            delattr(self.parent, 'name_window')

    def calculateDistances_thread(
        self,
        data_frame,
        time_column,
        names,
        ):
        """
        Performs distance calculations and outputs to file.
        """

        # Process strings

        names = [HeatMapPage.standardize_string(name) for name in
                 HeatMapPage.process_string_input(names)]
        print("names after processing:", names)

        outputWriter = csv.writer(self.outputFile)
        outputList = []
        outputList.append(time_column)
        for ind in range(len(names)):
            for ind2 in range(ind + 1, len(names)):
                outputList.append('Distance from ' + names[ind] + ' to '
                                   + names[ind2])
        outputWriter.writerow(outputList)
        outputList.clear()

        # Writes to csv if criteria is met.

        split_data = self.filter_by_time(data_frame, time_column)

        iterNames = []
        points = []
        if self.zcol_name != '':
            for (timestamp, time_data) in split_data:
                if type(timestamp) == tuple:
                    if time_column == 'Date' and len(str(timestamp)) \
                        > 10:
                        outputList.append(str(timestamp[0])[:10])
                    elif time_column == 'Time' and len(str(timestamp)) \
                        > 10:
                        outputList.append(str(timestamp[1]))
                    else:
                        outputList.append(str(timestamp[0])[:10] + ' '
                                + str(timestamp[1]))
                else:
                    if time_column == 'Date' and len(str(timestamp)) \
                        > 10:
                        outputList.append(str(timestamp)[:10])
                    elif time_column == 'Time' and len(str(timestamp)) \
                        > 10:
                        outputList.append(str(timestamp)[11:])
                    else:
                        outputList.append(timestamp)
                for (row_index, row) in time_data.iterrows():
                    name = \
                        HeatMapPage.standardize_string(row[self.name_column])
                    if name in names:
                        iterNames.append(name)
                        points.append((row[self.xcol_name],
                                row[self.ycol_name],
                                row[self.zcol_name]))

                if len(points) < 2:
                    outputList.clear()
                    iterNames.clear()
                    points.clear()
                else:
                    for ind in range(len(names)):
                        if names[ind] in iterNames:
                            for ind2 in range(ind + 1, len(names)):
                                if names[ind2] in iterNames:
                                    outputList.append(self.calculateDistance2Names3D(
                                        points[iterNames.index(names[ind])][0],
                                        points[iterNames.index(names[ind])][1],
                                        points[iterNames.index(names[ind])][2],
                                        points[iterNames.index(names[ind2])][0],
                                        points[iterNames.index(names[ind2])][1],
                                        points[iterNames.index(names[ind2])][2],
                                        ))
                                else:
                                    outputList.append('NA')
                        else:
                            for ind2 in range(ind + 1, len(names)):
                                outputList.append('NA')
                    outputWriter.writerow(outputList)
                    outputList.clear()
                    iterNames.clear()
                    points.clear()
        else:
            for (timestamp, time_data) in split_data:
                if type(timestamp) == tuple:
                    if time_column == 'Date' and len(str(timestamp)) \
                        > 10:
                        outputList.append(str(timestamp[0])[:10])
                    elif time_column == 'Time' and len(str(timestamp)) \
                        > 10:
                        outputList.append(str(timestamp[1]))
                    else:
                        outputList.append(str(timestamp[0])[:10] + ' '
                                + str(timestamp[1]))
                else:
                    if time_column == 'Date' and len(str(timestamp)) \
                        > 10:
                        outputList.append(str(timestamp)[:10])
                    elif time_column == 'Time' and len(str(timestamp)) \
                        > 10:
                        outputList.append(str(timestamp)[11:])
                    else:
                        outputList.append(timestamp)
                for (row_index, row) in time_data.iterrows():
                    name = \
                        HeatMapPage.standardize_string(row[self.name_column])
                    if name in names:
                        iterNames.append(name)
                        points.append((row[self.xcol_name],
                                row[self.ycol_name]))

                if len(points) < 2:
                    outputList.clear()
                    iterNames.clear()
                    points.clear()
                else:
                    for ind in range(len(names)):
                        if names[ind] in iterNames:
                            for ind2 in range(ind + 1, len(names)):
                                if names[ind2] in iterNames:
                                    outputList.append(self.calculateDistance2Names(points[iterNames.index(names[ind])][0],
        points[iterNames.index(names[ind])][1],
        points[iterNames.index(names[ind2])][0],
        points[iterNames.index(names[ind2])][1]))
                                else:
                                    outputList.append('NA')
                        else:
                            for ind2 in range(ind + 1, len(names)):
                                outputList.append('NA')
                    outputWriter.writerow(outputList)
                    outputList.clear()
                    iterNames.clear()
                    points.clear()

        # Closes file and destroys window when complete

        self.outputFile.close()

        # time.sleep(20)

        CalculationsCompleteMessage(self.filename)

        self.destroy()

    def calculateDistance2Names(
        self,
        x1,
        y1,
        x2,
        y2,
        ):
        """
        Performs 2D distance calculations using calibration ratios
        """

        if self.x_ratio != '' and self.y_ratio != '':
            return np.sqrt(np.square(x1 * self.x_ratio - x2
                           * self.x_ratio) + np.square(y1
                           * self.y_ratio - y2 * self.y_ratio))
        else:
            return np.sqrt(np.square(x1 - x2) + np.square(y1 - y2)) \
                * self.cal_ratio

    def calculateDistance2Names3D(
        self,
        x1,
        y1,
        z1,
        x2,
        y2,
        z2,
        ):
        """
        Performs 3D distance calculations using calibration ratios
        """

        if self.x_ratio != '' and self.y_ratio != '' and self.z_ratio \
            != '':
            return np.sqrt(np.square(x1 * self.x_ratio - x2
                           * self.x_ratio) + np.square(y1
                           * self.y_ratio - y2 * self.y_ratio)
                           + np.square(z1 * self.z_ratio - z2
                           * self.z_ratio))
        else:
            return np.sqrt(np.square(x1 - x2) + np.square(y1 - y2)
                           + np.square(z1 - z2)) * self.cal_ratio

    def filter_by_time(self, data_frame, time_column):
        """
        Handles use of Date + Time columns for calculations
        """

        if time_column == 'Date + Time':
            return data_frame.groupby(['Date', 'Time'])
        else:
            return data_frame.groupby([time_column])


class CalculationsCompleteMessage(tk.Toplevel):

    """
    Window that lets the user know calculations are complete for a file
    """

    def __init__(self, filename):
        tk.Toplevel.__init__(self)
        self.protocol('WM_DELETE_WINDOW', lambda : self.return_cancel())
        self.attributes('-topmost', 'true')

        self.title('Calculations Complete')

        tk.Toplevel.wm_geometry(self, '300x100')

        frm = tk.Frame(self, borderwidth=4, relief='raised')
        self.frame = frm
        self.frame.config(bg='white')
        frm.pack(fill='both', expand=True)

        self.enter_text = tk.Label(frm, text='The calcuations for '
                                   + filename + ' have finished')
        self.enter_text.pack(pady=10)

        b_ok = tk.Button(frm, text='OK')
        b_ok['command'] = lambda : self.destroy()
        b_ok.pack(side=tk.RIGHT, padx=4, pady=4)

    def return_cancel(self):
        self.destroy()


class SheetView(tk.Tk, tk.Toplevel):

    """
    Class that creates Spreadsheet window for interaction with the graph
    """

    def __init__(
        self,
        dataframe,
        options,
        controller,
        ):
        tk.Tk.__init__(self)
        menu = tk.Menu(self)
        self.config(menu=menu)
        self.original_df = dataframe
        subMenu = tk.Menu(menu, tearoff=0)
        subMenu.add_command(label='Save New Spreadsheet',
                            command=lambda : \
                            self.save_current_spreadsheet())
        subMenu.add_command(label='Sort Sheet Data By...',
                            command=lambda : \
                            self.handle_sort_spreadsheet(dataframe))
        subMenu.add_command(label='Filter Sheet Data By...',
                            command=lambda : \
                            self.handle_filter_selection(dataframe,
                            options))
        subMenu.add_command(label='Remove Filter...', command=lambda : \
                            self.handle_remove_filter(dataframe,
                            options))

        sortMenu = tk.Menu(subMenu, tearoff=0)

        # sortMenu.add_command(label=str(col), command = lambda : self.sort_spreadsheet(dataframe, col))

        # subMenu.add_cascade(label="Sort Sheet Data By...", menu=sortMenu)

        menu.add_cascade(label='File', menu=subMenu)
        self.protocol('WM_DELETE_WINDOW', lambda : \
                      self.close(controller))
        self.grid_columnconfigure(0, weight=1)
        self.grid_rowconfigure(0, weight=1)
        self.frame = tk.Frame(self)
        self.frame.grid_columnconfigure(0, weight=1)
        self.frame.grid_rowconfigure(0, weight=1)

        # dataframe = dataframe.fillna('')

        dftemp = dataframe.fillna('')
        self.current_df = dftemp
        self.sheet = Sheet(
            self.frame,
            data=dftemp.values.tolist(),
            headers=list(dataframe.columns),
            width=1440,
            height=810,
            column_width=200,
            )
        self.sheet.enable_bindings()
        self.sheet.popup_menu_add_command('Update plot',
                self.update_plot)
        self.frame.grid(row=0, column=0, sticky='nswe')
        self.sheet.grid(row=0, column=0, sticky='nswe')

    def handle_filter_selection(self, df, opts):
        """
        Opens new filter window if one is not already open.
        """

        if not hasattr(self, 'filterMenu'):
            self.filter_spreadsheet(df, opts)
        else:
            self.filterMenu.deiconify()

    def remove_filter(self, dataf, ops):
        """
    ....Removes an existing filter from the spreadsheet and updates accordingly.
    ...."""

        self.removeFilterMenu = tk.Toplevel(self)
        self.removeFilterMenu.protocol('WM_DELETE_WINDOW', lambda : \
                self.filter_del_return())
        self.removeFilterMenu.wm_title('Choose a Filter Name to Remove')
        currentLabel = tk.Label(self.removeFilterMenu,
                                text='Current Filters:',
                                font=('Helvetica', 10))
        currentLabel.pack(side='top', fill='x', pady=10)
        for filter_name in ops['filters']:
            filterLabel = tk.Label(self.removeFilterMenu,
                                   text='{} -> {}'.format(filter_name,
                                   ops['filters'][filter_name]),
                                   font=('Helvetica', 8))
            filterLabel.pack(side='top', fill='x', pady=10)
        label1 = tk.Label(self.removeFilterMenu,
                          text='Please choose a filter name',
                          font=('Helvetica', 10))
        label1.pack(side='top', fill='x', pady=10)
        tkvar = StringVar(self.removeFilterMenu)
        titles = ops['filters'].keys()
        options = list(titles)

        if len(options) != 0:
            drop = tk.OptionMenu(self.removeFilterMenu, tkvar, *options)
            drop.pack()
            button1 = tk.Button(self.removeFilterMenu, text='Submit',
                                command=lambda : \
                                self.submit_remove_choice(tkvar, ops))
            button1.pack()
            button2 = tk.Button(self.removeFilterMenu, text='Cancel',
                                command=self.filter_del_return)
            button2.pack()
            self.removeFilterMenu.mainloop()

    def filter_spreadsheet(self, df, opts):
        """
        Filters spreadhsheet based on input column and value.
        """

        self.filterMenu = tk.Toplevel(self)
        self.filterMenu.protocol('WM_DELETE_WINDOW', lambda : \
                                 self.filter_return())
        self.filterMenu.wm_title('Choose a Column and Value To Filter By'
                                 )
        currentLabel = tk.Label(self.filterMenu, text='Current Filters:'
                                , font=('Helvetica', 10))
        currentLabel.pack(side='top', fill='x', pady=10)

        for filter_name in opts['filters']:
            filterLabel = tk.Label(self.filterMenu,
                                   text='{} -> {}'.format(filter_name,
                                   opts['filters'][filter_name]),
                                   font=('Helvetica', 8))
            filterLabel.pack(side='top', fill='x', pady=10)

        label1 = tk.Label(self.filterMenu,
                          text='Please choose a column name',
                          font=('Helvetica', 10))
        label1.pack(side='top', fill='x', pady=10)

        tkvar = StringVar(self.filterMenu)
        options = list(df.columns)
        drop = tk.OptionMenu(self.filterMenu, tkvar, *options)
        drop.pack()

        label2 = tk.Label(self.filterMenu,
                          text='Please enter a value to filter by',
                          font=('Helvetica', 10))
        label2.pack(side='top', fill='x', pady=10)

        data_val_box = tk.Text(self.filterMenu, height=1, width=32)
        data_val_box.pack(side='top', pady=10)

        button1 = tk.Button(self.filterMenu, text='Submit',
                            command=lambda : \
                            self.submit_filter_choice(tkvar,
                            data_val_box, opts))
        button1.pack()
        button2 = tk.Button(self.filterMenu, text='Cancel',
                            command=self.filter_return)
        button2.pack()
        self.filterMenu.mainloop()

    def filter_del_return(self):
        self.removeFilterMenu.destroy()
        delattr(self, 'removeFilterMenu')

    def filter_return(self):
        """
        Handles closing of filter window
        """

        self.filterMenu.destroy()
        delattr(self, 'filterMenu')

    def update_sheet(self, dataframe):
        """
        Fills empty entries in data with blanks
        """

        dataframe = dataframe.fillna('')
        self.sheet.set_sheet_data(data=dataframe.values.tolist())
        self.sheet.refresh()

    def submit_remove_choice(self, cvar, opts):
        """
    ....Handles submission of filter removal and only returning original data with matching values
    ...."""

        opts['filters'].pop(str(cvar.get()), None)

        dataframe = self.original_df
        for k in opts['filters']:
            if opts['filters'][k].isnumeric():
                num = int(opts['filters'][k])
                dataframe = dataframe[dataframe[k] == num]
            else:
                dataframe = dataframe[dataframe[k] == str(opts['filters'
                        ][k])]

        self.current_df = dataframe
        self.update_sheet(self.current_df)
        self.removeFilterMenu.destroy()
        delattr(self, 'removeFilterMenu')

    def handle_remove_filter(self, dataframe, options):
        """
        Opens sort window if it's not already open
        """

        if not hasattr(self, 'removeFilterMenu'):

            # self.sort_spreadsheet(dataframe)

            self.remove_filter(dataframe, options)
        else:
            self.removeFilterMenu.deiconify()

    def remove_filter_return(self):
        """
        Handles closing of sort window
        """

        self.removeFilterMenu.destroy()
        delattr(self, 'removeFilterMenu')

    def submit_filter_choice(
        self,
        cvar,
        textbox,
        ops,
        ):
        """
        Handles submission of filter and only keeping data with matching values
        """

        dataframe = self.current_df
        column_choice = cvar.get()
        raw = textbox.get('1.0', 'end-1c')
        allowed = HeatMapPage.process_string_input(raw)

        # Filter data

        if dataframe[column_choice].dtypes == 'object':
            self.current_df = \
                dataframe.loc[dataframe[column_choice].astype(str).apply(HeatMapPage.standardize_string,
                              1).isin(allowed)]
        else:
            self.current_df = \
                dataframe.loc[dataframe[column_choice].isin(allowed)]

        # Update GUI and variables

        ops['filters'][column_choice] = ', '.join([str(val) for val in
                allowed])
        self.update_sheet(self.current_df)
        self.filterMenu.destroy()
        delattr(self, 'filterMenu')

    def handle_sort_spreadsheet(self, dataframe):
        """
        Opens sort window if it's not already open
        """

        if not hasattr(self, 'sortMenu'):
            self.sort_spreadsheet(dataframe)
        else:
            self.sortMenu.deiconify()

    def submit_sort_choice(self, svar):
        """
        Handles submission of sorting and sorts the dataframe accordingly
        """

        objtype = ''
        for i in self.current_df[svar.get()]:
            if i is not None and i != '':
                objtype = type(i)
                break

        # print(objtype)

        if objtype == float or objtype == int:
            self.current_df[svar.get()] = \
                pd.to_numeric(self.current_df[svar.get()],
                              errors='coerce')
        elif objtype == pd.Timestamp:
            self.current_df[svar.get()] = \
                pd.to_datetime(self.current_df[svar.get()],
                               errors='coerce')
        dfs = self.current_df.sort_values(by=[svar.get()])
        self.update_sheet(dfs)
        self.current_df = dfs
        self.sort_return()

    def sort_return(self):
        """
        Handles closing of sort window
        """

        self.sortMenu.destroy()
        delattr(self, 'sortMenu')

    def sort_spreadsheet(self, df):
        """
        Creats window for sorting by a certain column
        """

        self.sortMenu = tk.Toplevel(self)
        self.sortMenu.protocol('WM_DELETE_WINDOW', lambda : \
                               self.sort_return())
        self.sortMenu.wm_title('Choose a Column To Sort By')
        label = tk.Label(self.sortMenu,
                         text='Please choose a column name from the list below'
                         , font=('Helvetica', 10))
        label.pack(side='top', fill='x', pady=10)

        tkvar = StringVar(self.sortMenu)

        options = list(df.columns)

        drop = tk.OptionMenu(self.sortMenu, tkvar, *options)

        drop.pack()

        button1 = tk.Button(self.sortMenu, text='Submit',
                            command=lambda : \
                            self.submit_sort_choice(tkvar))
        button1.pack()

        button2 = tk.Button(self.sortMenu, text='Cancel',
                            command=self.sort_return)
        button2.pack()

        self.sortMenu.mainloop()

    def save_current_spreadsheet(self):
        """
        Opens file explorer to choose where to save and export the current spreadsheet.
        """

        types = [('CSV file (.csv)', '*.csv'), ('Excel file (.xlsx)',
                 '*.xlsx')]
        file = filedialog.asksaveasfile(filetypes=types,
                defaultextension='*.csv', initialfile='.csv')

        if file:
            (path, ext) = os.path.splitext(file.name)

            if ext == '.csv':
                self.current_df.to_csv(path_or_buf=file.name,
                        index=False)
            else:
                self.current_df.to_excel(excel_writer=file.name,
                        index=False)

    def update_plot(self):
        """
        Updates the graph based on the spreadsheet
        """

        app.update_graph(self.sheet)

    def close(self, controller):
        """
        Closes sheet
        """

        delattr(controller, 'sheet')
        self.destroy()


app = ZooMapper()
app.mainloop()
