from openpyxl import load_workbook

file_path = 'Sample Data for 3D Distances.xlsx'

wb = load_workbook(file_path)
sheet = wb.active
print(f"Total number of row in the present sheet is {sheet.max_row}")
rows = sheet.max_row
ws = wb['SharkFacility4']

list_with_values = []
for cell in ws[1]: # grabbing the names of all of our columns
    list_with_values.append(cell.value)
    
indexes_to_invert = []
i = 1
for cell in ws[1]: # grabbing index of each col we want to change
    if (cell.value == "DepthM"):
        print("D", i)
        indexes_to_invert.append(i)
    if (cell.value == "Space Use Coordinate Y"):
        print("Y", i)
        indexes_to_invert.append(i)
    if (cell.value == "Space Use Coordinate X"):
        print("X", i)
        indexes_to_invert.append(i)
    i+=1
 
print(sheet.max_row)
for index in indexes_to_invert: # going theough the indexes whose values we want to invert
    print(index)
    for row in range(rows):
        row+=1
        if row != 1:
            cell = sheet.cell(row, index)
            inverted_num = cell.value * -1
            cell.value = inverted_num            
        
    
wb.save("inverted.xlsx") # saving the modifications we made to the column
wb.close